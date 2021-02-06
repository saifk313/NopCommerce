import openpyxl


def getRows(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    return sheet.max_row


def getColumns(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    return sheet.max_column


def readData(path, sheet_name, row_no, column_no):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    return sheet.cell(row=row_no, column=column_no).value


def writeData(path, sheet_name, row_no, column_no, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.cell(row=row_no, column=column_no).value = data
    workbook.save(path)
